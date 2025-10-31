import os
import re
import json
import base64
import fitz
import tempfile
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
from openai import OpenAI

# ========== CONFIG ==========
load_dotenv()

# Try Streamlit secrets first, then fall back to .env file
OPENAI_API_KEY = st.secrets.get("OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY")

if not OPENAI_API_KEY:
    st.error("❌ OPENAI_API_KEY not found in Streamlit secrets or .env file")
    st.stop()


client = OpenAI(api_key=OPENAI_API_KEY)

UPLOAD_FOLDER = "uploads"
DOWNLOAD_FOLDER = "downloads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

# ========== HELPERS ==========
def parse_page_list_single(range_str):
    if "-" in range_str:
        a, b = range_str.split("-", 1)
        return list(range(int(a), int(b) + 1))
    return [int(range_str)]

def parse_pages_input(pages_input):
    pages = []
    for part in re.split(r"[,\s]+", pages_input.strip()):
        if not part:
            continue
        pages.extend(parse_page_list_single(part))
    return sorted(set(pages))

def split_into_blocks(all_pages):
    if not all_pages:
        return [], [], []
    if len(all_pages) == 1:
        return [all_pages[0]], [], []
    subject = [all_pages[0]]
    cost = [all_pages[-1]]
    sales = [p for p in all_pages if p not in subject + cost]
    return subject, sales, cost

def pdf_pages_to_images(pdf_path, pages, dpi=250):
    doc = fitz.open(pdf_path)
    out_paths = []
    for p in pages:
        if 1 <= p <= len(doc):
            out = os.path.join(UPLOAD_FOLDER, f"page_{p}.png")
            doc.load_page(p - 1).get_pixmap(dpi=dpi).save(out)
            out_paths.append(out)
    doc.close()
    return out_paths

def encode_image_b64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def clean_json_output(raw_text):
    text = raw_text.strip()
    text = re.sub(r"^```(json)?", "", text)
    text = re.sub(r"```$", "", text)
    match = re.search(r"\{[\s\S]*\}", text)
    candidate = match.group(0) if match else text
    try:
        return json.loads(candidate)
    except Exception:
        return {"error": "Invalid JSON returned from model", "raw_output": raw_text}

def build_image_messages(img_paths, prompt_text):
    imgs = []
    for p in img_paths:
        img_b64 = encode_image_b64(p)
        imgs.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}})
    return [
        {"role": "system", "content": "You are a structured data extraction model. Return only valid JSON."},
        {"role": "user", "content": [{"type": "text", "text": prompt_text}, *imgs]},
    ]

def merge_dicts(*dicts):
    merged = {}
    for d in dicts:
        if not isinstance(d, dict):
            continue
        for k, v in d.items():
            if k not in merged:
                merged[k] = v
            elif isinstance(merged[k], dict) and isinstance(v, dict):
                merged[k].update(v)
    return merged

# ========== PROMPTS ==========
PROMPT_SUBJECT = """
Extract all visible text-based fields for the SUBJECT block and related sections.
Pay extra attention to small or faint labels like HOA Frequency and One-Unit Housing Trends.

ONE-UNIT HOUSING TRENDS:
["Property Values","Demand/Supply","Marketing Time"]
(These fields are often small and appear near the top of the page. Must extract even if the text is faint.)

SUBJECT:
["Property Address","County","City","State","Zip Code","Assessor's Parcel #","R.E. Taxes $","Neighborhood Name","HOA $","R.E Tax Year","HOA Frequency","Pool"]

ONE-UNIT HOUSING:
["Price $ (000) - Low","Price $ (000) - High","Price $ (000) - Pred","Age (Yrs) - Low","Age (Yrs) - High","Age (Yrs) - Pred"]

SITE:
["Area","Specific Zoning Classification","Zoning Description","FEMA Flood Zone","FEMA Map #","FEMA Map Date"]

IMPROVEMENTS:
["Year Built","Garage"]

Special field instructions:
- "State" must be exactly the 2-letter postal abbreviation (e.g., WI, IL, CA). If unreadable or not visible, set to " ".
- Extract only marked checkbox value for "HOA Frequency" (e.g., per month, per year). If no box is marked, set to " ".
- "Zoning Description" must be zoning-related text like "Residential/Planned Development", "Apartment House", " "(not "See" or "Addendum").
- "Specific Zoning Classification" should be codes like "R-1", "R-2", "C-3", "A-1", "PD", " ".
- Visible text only. Return valid JSON with top-level keys: SUBJECT, ONE_UNIT_HOUSING_TRENDS, ONE_UNIT_HOUSING, SITE, IMPROVEMENTS.
- Missing fields → " ".
"""

PROMPT_SALES = """
Extract fields for SALES COMPARISON, SALES HISTORY, and RECONCILIATION.

SALES_COMPARISON_APPROACH (Subject + Comparable Sale #1–#9):
["Address","Proximity","Sale Price","Data Source","Verification Source","Date of Sale","Location","Site","View","Quality","Condition","Total","Bdrms","Baths","GLA","Basement","Pool","Garage", "Date of Prior Sale/Transfer","Price of Prior Sale/Transfer","Data Source(s)","Effective Date of Data Source(s)"]

RECONCILIATION:
["Sales Comparison Approach","Cost Approach","Income Approach","Value","Effective Date","Opinion of Site value","Total Estimate of Cost-New","Depreciation"]

Rules:
- Include only "Subject" and "Comparable Sale #1" through "Comparable Sale #9".
- Ignore any additional comparables (#10 or beyond).
- Each column (Subject, #1–#9) is separate.
- Blank or placeholder → " ".
- Return valid JSON with top-level keys: SALES_COMPARISON_APPROACH, SALES_HISTORY, RECONCILIATION.
- Keep "Total", "Bdrms", and "Baths" as separate numeric/string fields.
- If some values missing in the "Subject" do not extract from the "Comparables".
- Ignore extracting "Proximity","Sale Price","Data Source","Verification Source","Date of Sale" for the "Subject" entry. Extract for comparables only.
- Dont extract "Date of Prior Sale/Transfer","Price of Prior Sale/Transfer","Data Source(s)","Effective Date of Data Source(s)" the Subject value for Comparables. Extract Subject and Comparables seperately.
- For "Sales Comparison Approach","Cost Approach","Income Approach" extracttheir values from RECONCILIATION section they are inndicate after $ sign as a numeric value.
- For "Value" extract the value indicated after "As Is Value" in RECONCILIATION section.
-For "Effective Date" extract the date indicated after "Effective Date" in RECONCILIATION section.
"""

PROMPT_COST = """
Extract fields for COST APPROACH and RECONCILIATION.

COST_APPROACH:
["Opinion of Site value","Total Estimate of Cost-New","Depreciation"]



Rules:
- Visible text only.
- Blank → " ".
- Return valid JSON with top-level keys COST_APPROACH.

"""

PROMPT_SITE = """
"""

# ========== STREAMLIT UI ==========
st.set_page_config(page_title="PARK Extractor", page_icon="🧠", layout="wide")

st.title("📄 PARK Appraisal Report Extractor")
st.caption("Specialized for **Form 1004** appraisal document extraction")

tab1, tab2 = st.tabs(["1️⃣ Extract JSON from PDF", "2️⃣ Fill Excel from JSON"])

# --- TAB 1: PDF → JSON Extraction ---
with tab1:
    st.subheader("Extract structured data from Form 1004 PDF")

    pdf_file = st.file_uploader("Upload your Form 1004 PDF", type=["pdf"])
    pages_input = st.text_input("Enter Page Range (e.g. 2-8 or 2-5,8):", "2-8")

    if st.button("🔍 Extract Data"):
        if not pdf_file:
            st.warning("Please upload a PDF first.")
        else:
            pdf_path = os.path.join(UPLOAD_FOLDER, pdf_file.name)
            with open(pdf_path, "wb") as f:
                f.write(pdf_file.read())

            all_pages = parse_pages_input(pages_input)
            subj, sales, cost = split_into_blocks(all_pages)
            base_name = os.path.splitext(pdf_file.name)[0]

            with st.spinner("Extracting from pages..."):
                results = {"blocks": {"subject": subj, "sales": sales, "cost": cost}, "raw": {}, "merged": {}}

                # SUBJECT
                imgs = pdf_pages_to_images(pdf_path, subj)
                resp = client.chat.completions.create(model="gpt-4.1-mini", temperature=0,
                                                      messages=build_image_messages(imgs, PROMPT_SUBJECT))
                results["raw"]["subject"] = clean_json_output(resp.choices[0].message.content)

                # SALES
                imgs = pdf_pages_to_images(pdf_path, sales)
                resp = client.chat.completions.create(model="gpt-4.1-mini", temperature=0,
                                                      messages=build_image_messages(imgs, PROMPT_SALES))
                results["raw"]["sales"] = clean_json_output(resp.choices[0].message.content)

                # COST
                imgs = pdf_pages_to_images(pdf_path, cost)
                resp = client.chat.completions.create(model="gpt-4.1-mini", temperature=0,
                                                      messages=build_image_messages(imgs, PROMPT_COST))
                results["raw"]["cost"] = clean_json_output(resp.choices[0].message.content)

                results["merged"] = merge_dicts(results["raw"].get("subject"),
                                                results["raw"].get("sales"),
                                                results["raw"].get("cost"))

                out_path = os.path.join(DOWNLOAD_FOLDER, f"{base_name}_extracted.json")
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(results, f, indent=2)

                st.success("✅ Extraction Complete!")
                st.json(results["merged"])
                st.download_button("⬇️ Download JSON File",
                                   data=json.dumps(results, indent=2),
                                   file_name=f"{base_name}_extracted.json",
                                   mime="application/json")

# --- TAB 2: JSON → Excel Filling ---
with tab2:
    st.subheader("Fill Excel Template from JSON")

    excel_file = st.file_uploader("Upload Excel Template", type=["xlsx"])
    json_file = st.file_uploader("Upload Extracted JSON", type=["json"])

    if st.button("📊 Fill Excel"):
        if not excel_file or not json_file:
            st.warning("Please upload both Excel and JSON files.")
        else:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel:
                temp_excel.write(excel_file.read())
                temp_excel_path = temp_excel.name

            wb = load_workbook(temp_excel_path)
            ws = wb.active

            merged_json = json.load(json_file).get("merged", {})

            # Reuse all functions from your original Flask JSON-to-Excel code
            # (flatten_json, canonical_field, canonical_section, detect_sections, etc.)
            # Include same logic below:

            # --- COPY YOUR JSON→EXCEL FUNCTIONS (no edits) ---
            import re

            ROW_SECTION = 2
            ROW_FIELDS = 3
            FIRST_DATA_ROW = 4

            def norm(s): return " ".join(str(s).replace("\n"," ").strip().replace("# ","#").split()).upper() if s else ""

            FIELD_SYNONYMS = {
                "BDRMS":"BEDROOMS",
                "BATH(S)":"BATHS",
                "QUALITY":"QUALITY OF CONSTRUCTION",
                "QUALITY OF CONST":"QUALITY OF CONSTRUCTION",
                "PROPERTY ADDRESS":"PROPERTY ADDRESS",
                "ADDRESS":"ADDRESS",
                "LOT SIZE":"SITE",
                "AREA":"SITE"
            }

            FIELD_ONLY_SET = {"OPINION OF SITE VALUE","TOTAL ESTIMATE OF COST-NEW","DEPRECIATION","R.E. TAX YEAR","HOA FREQUENCY","POOL"}

            SALES_HISTORY_COMP_FIELDS = {"DATE OF PRIOR SALE/TRANSFER","PRICE OF PRIOR SALE/TRANSFER","DATA SOURCE(S)","EFFECTIVE DATE OF DATA SOURCE(S)"}

            def canonical_field(s):
                s=norm(s)
                return FIELD_SYNONYMS.get(s, s)

            def canonical_section(s):
                s=norm(s)
                m=re.match(r"COMPARA?BE\s*(\d+)", s)
                if m: return f"COMPARABLE SALE #{m.group(1)}"
                if s.startswith("COMPARABLE SALE #"):
                    return re.sub(r"COMPARABLE SALE #\s*(\d+)", r"COMPARABLE SALE #\1", s)
                if "ONE-UNIT HOUSING TRENDS" in s: return "ONE_UNIT_HOUSING_TRENDS"
                if "ONE-UNIT HOUSING" in s: return "ONE_UNIT_HOUSING"
                if "RECONCILIATION" in s: return "RECONCILIATION"
                if "COST APPROACH" in s: return "COST_APPROACH"
                if "SALES" in s and "HISTORY" in s: return "SALES_HISTORY"
                if s == "SUBJECT": return "SUBJECT"
                return s

            def flatten_json(merged):
                flat={}
                def walk(path,obj):
                    if isinstance(obj, dict):
                        for k,v in obj.items():
                            walk(path+[k], v)
                    else:
                        raw_sections=path[:-1]
                        raw_field=path[-1]
                        field_c=canonical_field(raw_field)
                        section_c=""
                        for p in reversed(raw_sections):
                            sec=canonical_section(p)
                            if sec:
                                section_c=sec
                                break
                        val=str(obj)
                        if section_c and field_c:
                            flat[f"{section_c}.{field_c}"]=val
                        if field_c in FIELD_ONLY_SET:
                            flat[field_c]=val
                walk([], merged)
                return flat

            def detect_sections():
                max_col=ws.max_column
                raw=[]
                last=""
                for col in range(1,max_col+1):
                    sec=canonical_section(ws.cell(row=ROW_SECTION,column=col).value)
                    if sec: last=sec
                    raw.append(last)
                comp1_start = next((i for i,x in enumerate(raw,1) if x=="COMPARABLE SALE #1"), None)
                final={}
                for col in range(1,max_col+1):
                    sec=raw[col-1]
                    field=canonical_field(ws.cell(row=ROW_FIELDS,column=col).value)
                    if not sec:
                        if comp1_start and col<comp1_start:
                            sec="SUBJECT"
                        elif field in SALES_HISTORY_COMP_FIELDS:
                            sec="SALES_HISTORY_SUBJECT"
                    final[col]=sec
                return final

            def pick_value(section, field, flat):
                if field in FIELD_ONLY_SET:
                    return flat.get(field,"")
                if field in SALES_HISTORY_COMP_FIELDS and section.startswith("COMPARABLE SALE #"):
                    return flat.get(f"{section}.{field}","")
                if section.startswith("COMPARABLE SALE #"):
                    return flat.get(f"{section}.{field}","")
                if section=="SUBJECT":
                    return flat.get(f"SUBJECT.{field}","")
                if section=="SALES_HISTORY_SUBJECT":
                    return flat.get(f"SALES_HISTORY.Subject.{field}","")
                return flat.get(f"{section}.{field}","")

            flat = flatten_json(merged_json)
            secmap = detect_sections()

            colmap={}
            for col in range(1, ws.max_column+1):
                sec=secmap[col]
                field=canonical_field(ws.cell(row=ROW_FIELDS,column=col).value)
                if sec and field:
                    colmap.setdefault((sec,field), col)

            new_row=max(ws.max_row+1, FIRST_DATA_ROW)
            filled=0
            for (sec,field),col in colmap.items():
                current=ws.cell(row=new_row, column=col).value
                if current not in (None,""," "):
                    continue
                val=pick_value(sec,field,flat)
                if val:
                    ws.cell(row=new_row, column=col).value=val
                    filled+=1

            out_xlsx = os.path.join(DOWNLOAD_FOLDER, "1004_updated.xlsx")
            wb.save(out_xlsx)

            st.success(f"✅ Added Property — {filled} fields populated!")
            with open(out_xlsx, "rb") as f:
                st.download_button("⬇️ Download Updated Excel", f, file_name="1004_updated.xlsx")







